"""
CAD bulk file loader.

Reads a county CAD export (CSV or XLSX), normalizes column names using the
county's column_map, and yields standardized CadParcel dicts ready for the
database writer.

Usage:
    config = COUNTY_CONFIGS["travis"]
    for parcel in load_cad_file(config, "/path/to/export.csv"):
        upsert_parcel(parcel)
"""

import csv
import logging
from pathlib import Path
from typing import Iterator, Optional

from ingestion.cad.counties import CadCountyConfig

logger = logging.getLogger(__name__)


def _clean(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = value.strip()
    return value if value else None


def _to_int(value: Optional[str]) -> Optional[int]:
    # Strip commas ("1,800") and truncate trailing .0 ("1800.0") before parsing.
    try:
        cleaned = value.replace(",", "").strip() if value else ""
        return int(float(cleaned)) if cleaned else None
    except ValueError:
        return None


def _to_float(value: Optional[str]) -> Optional[float]:
    try:
        cleaned = value.replace(",", "").strip() if value else ""
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def load_cad_file(config: CadCountyConfig, file_path: str | Path) -> Iterator[dict]:
    """
    Parse a CAD export file and yield normalized parcel dicts.

    Yields dicts with keys matching the `properties` table columns:
        apn, owner_name, address_raw, city, county, state, zip_code,
        land_value, improvement_value, total_cad_value,
        sqft, year_built, bedrooms, bathrooms
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        yield from _load_csv(config, path)
    elif suffix in (".xlsx", ".xls"):
        yield from _load_excel(config, path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}")


def _load_csv(config: CadCountyConfig, path: Path) -> Iterator[dict]:
    inv_map = {v: k for k, v in config.column_map.items()}

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for row_num, row in enumerate(reader, start=2):
            try:
                yield _normalize_row(row, inv_map, config)
            except Exception as exc:
                logger.warning("Row %d skipped — %s", row_num, exc)


def _load_excel(config: CadCountyConfig, path: Path) -> Iterator[dict]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Install openpyxl: pip install openpyxl") from exc

    inv_map = {v: k for k, v in config.column_map.items()}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {h: (str(v) if v is not None else None) for h, v in zip(headers, row)}
        try:
            yield _normalize_row(raw, inv_map, config)
        except Exception as exc:
            logger.warning("Row %d skipped — %s", row_num, exc)

    wb.close()


def _normalize_row(row: dict, inv_map: dict, config: CadCountyConfig) -> dict:
    """Map raw county column names to standard field names."""
    std: dict = {}
    for raw_col, value in row.items():
        std_col = inv_map.get(raw_col)
        if std_col:
            std[std_col] = value

    apn = _clean(std.get("apn"))
    if not apn:
        raise ValueError("Missing APN — row skipped")

    return {
        "apn": apn,
        "county": config.name.lower(),
        "state": "TX",
        "owner_name": _clean(std.get("owner_name")),
        "address_raw": _clean(std.get("address_raw")),
        "city": _clean(std.get("city")),
        "zip_code": _clean(std.get("zip_code")),
        "land_value": _to_float(std.get("land_value")),
        "improvement_value": _to_float(std.get("improvement_value")),
        "total_cad_value": _to_float(std.get("total_value")),
        "sqft": _to_int(std.get("sqft")),
        "year_built": _to_int(std.get("year_built")),
        "bedrooms": _to_int(std.get("bedrooms")),
        "bathrooms": _to_float(std.get("bathrooms")),
    }
